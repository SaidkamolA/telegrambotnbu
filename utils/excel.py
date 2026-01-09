import os
import tempfile
from collections import Counter
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from survey.questions import SURVEY, SECTION_TITLES


THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_DARK = PatternFill("solid", fgColor="111827")
FILL_LIGHT = PatternFill("solid", fgColor="F3F4F6")
FILL_SECTION = PatternFill("solid", fgColor="0B1220")
FILL_Q = PatternFill("solid", fgColor="E5E7EB")
FILL_STRIPE = PatternFill("solid", fgColor="F8FAFC")

FONT_TITLE = Font(bold=True, size=14, color="FFFFFF")
FONT_SUB = Font(bold=True, size=10, color="111827")
FONT_SEC = Font(bold=True, size=11, color="FFFFFF")
FONT_Q = Font(bold=True, size=10, color="111827")
FONT_H = Font(bold=True, size=10, color="FFFFFF")
FONT_CELL = Font(size=10, color="111827")


def autosize_columns(ws, min_w=10, max_w=70):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), max_w)


def _section_title_for_qnum(q_num: int, lang: str) -> str:
    last = None
    for start_q, titles in SECTION_TITLES.items():
        if q_num >= start_q:
            last = titles.get(lang)
    return last or ""


def _fmt_percent(x: float) -> str:
    return f"{x:.2f}".replace(".", ",")


def _style_row(ws, row, fill=None, font=None, align=None, border=True):
    for cell in ws[row]:
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if border:
            cell.border = BORDER


def _create_dashboard_top(ws, totals: Counter, stats: Dict[str, Dict[str, Counter]], lang: str):
    title = "DASHBOARD_TOP" if lang == "ru" else "DASHBOARD_TOP_UZ"
    ws.title = title

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    total_all = int(totals.get("all", 0))
    total_ru = int(totals.get("ru", 0))
    total_uz = int(totals.get("uz", 0))
    denom = max(total_all, 1)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    # ✅ FIX: no NBU, just "Результаты"
    c.value = "Результаты (TOP) | I–VI (Q1–Q30) [RU]" if lang == "ru" else "Natijalar (TOP) | I–VI (Q1–Q30) [UZ]"
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value = (
        f"Total: {total_all} | RU: {total_ru} | UZ: {total_uz} | Share % from total respondents"
        if lang == "ru"
        else f"Jami: {total_all} | RU: {total_ru} | UZ: {total_uz} | Ulush % (jami respondentlardan)"
    )
    c2.fill = FILL_LIGHT
    c2.font = FONT_SUB
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    row = 4

    def add_section_header(title_text: str):
        nonlocal row

        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = title_text
        cell.fill = FILL_SECTION
        cell.font = FONT_SEC
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        headers = ["Q#", "Question (RU)", "Top option (RU)", "Count", "Share %", "Type"] if lang == "ru" else ["Q#", "Savol (UZ)", "Top javob (UZ)", "Soni", "Ulush %", "Turi"]
        ws.append(headers)
        _style_row(ws, row, fill=FILL_DARK, font=FONT_H, align=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[row].height = 18
        row += 1

    current_section = None

    for i, q in enumerate(SURVEY):
        q_num = i + 1

        sec = _section_title_for_qnum(q_num, lang)
        if sec != current_section:
            current_section = sec
            add_section_header(current_section)

        counter = stats.get("all", {}).get(q.key, Counter())
        q_text = q.text_ru if lang == "ru" else q.text_uz
        opts = q.options_ru if lang == "ru" else q.options_uz

        if counter:
            top_idx, top_cnt = counter.most_common(1)[0]
            top_opt = opts[top_idx] if top_idx < len(opts) else f"option[{top_idx}]"
            share_str = _fmt_percent((top_cnt / max(int(totals.get("all", 0)), 1)) * 100)
            cnt_val = int(top_cnt)
        else:
            top_opt = ""
            share_str = _fmt_percent(0.0)
            cnt_val = 0

        ws.append([q_num, q_text, top_opt, cnt_val, share_str, "single"])
        _style_row(ws, row, font=FONT_CELL, align=Alignment(vertical="center"))

        if (q_num % 2) == 0:
            for col in "ABCDEF":
                ws[f"{col}{row}"].fill = FILL_STRIPE

        row += 1

    for r in ws.iter_rows(min_row=1, max_row=row - 1, min_col=1, max_col=6):
        for cell in r:
            cell.border = BORDER


def _create_full_sheet_simple(wb: Workbook, title: str, lang_key: str, stats: Dict[str, Dict[str, Counter]]):
    ws = wb.create_sheet(title=title)
    ws.append(["Q#", "Question (RU)", "Option (RU)", "Option (UZ)", "Count"])

    for cell in ws[1]:
        cell.fill = FILL_DARK
        cell.font = FONT_H
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 18

    row = 2
    for i, q in enumerate(SURVEY):
        q_num = i + 1
        counter = stats.get(lang_key, {}).get(q.key, Counter())
        max_opts = max(len(q.options_ru), len(q.options_uz))

        for idx in range(max_opts):
            opt_ru = q.options_ru[idx] if idx < len(q.options_ru) else ""
            opt_uz = q.options_uz[idx] if idx < len(q.options_uz) else ""
            cnt = int(counter.get(idx, 0))

            ws.append([q_num, q.text_ru, opt_ru, opt_uz, cnt])
            for cell in ws[row]:
                cell.font = FONT_CELL
                cell.alignment = Alignment(vertical="center")
                cell.border = BORDER
            row += 1

    autosize_columns(ws)


def _create_all_pretty(wb: Workbook, totals: Counter, stats: Dict[str, Dict[str, Counter]]):
    ws = wb.create_sheet("ALL_PRETTY")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    total_all = int(totals.get("all", 0))

    # title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    # ✅ FIX: no NBU, just "Результаты"
    t.value = "Результаты | I–VI (Q1–Q30)"
    t.fill = FILL_DARK
    t.font = FONT_TITLE
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value = f"Total respondents: {total_all} | Percentages per question are calculated from answered count of that question"
    s.fill = FILL_LIGHT
    s.font = FONT_SUB
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    row = 4
    current_section = None

    def add_section(title_uz: str, title_ru: str):
        nonlocal row
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = f"{title_uz} / {title_ru}"
        c.fill = FILL_SECTION
        c.font = FONT_SEC
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    def add_question_header(q_num: int, q_ru: str, q_uz: str):
        nonlocal row
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value = f"{q_num}) {q_ru}  |  {q_uz}"
        c.fill = FILL_Q
        c.font = FONT_Q
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        ws.append(["№", "Option (RU)", "Option (UZ)", "Count", "Share %"])
        _style_row(ws, row, fill=FILL_DARK, font=FONT_H, align=Alignment(horizontal="center", vertical="center"))
        ws.row_dimensions[row].height = 18
        row += 1

    for i, q in enumerate(SURVEY):
        q_num = i + 1

        sec_ru = _section_title_for_qnum(q_num, "ru")
        sec_uz = _section_title_for_qnum(q_num, "uz")
        sec_mix = f"{sec_uz} / {sec_ru}"

        if sec_mix != current_section:
            current_section = sec_mix
            add_section(sec_uz, sec_ru)

        add_question_header(q_num, q.text_ru, q.text_uz)

        counter = stats.get("all", {}).get(q.key, Counter())
        answered = sum(counter.values())
        denom_q = max(answered, 1)

        max_opts = max(len(q.options_ru), len(q.options_uz))
        for idx in range(max_opts):
            opt_ru = q.options_ru[idx] if idx < len(q.options_ru) else ""
            opt_uz = q.options_uz[idx] if idx < len(q.options_uz) else ""
            cnt = int(counter.get(idx, 0))
            share = (cnt / denom_q) * 100 if answered else 0.0

            ws.append([idx + 1, opt_ru, opt_uz, cnt, _fmt_percent(share)])

            _style_row(ws, row, font=FONT_CELL, align=Alignment(vertical="center"))
            if (idx % 2) == 1:
                for col in "ABCDE":
                    ws[f"{col}{row}"].fill = FILL_STRIPE
            row += 1

        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"].value = "Answered (total for this question):"
        ws[f"A{row}"].font = Font(bold=True, size=10, color="111827")
        ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"A{row}"].border = BORDER

        ws[f"D{row}"].value = answered
        ws[f"D{row}"].font = Font(bold=True, size=10, color="111827")
        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{row}"].border = BORDER

        ws[f"E{row}"].value = ""
        ws[f"E{row}"].border = BORDER
        row += 2

    for r in ws.iter_rows(min_row=1, max_row=row - 1, min_col=1, max_col=5):
        for cell in r:
            if cell.border is None:
                cell.border = BORDER


def build_excel_stats(totals: Counter, stats: Dict[str, Dict[str, Counter]]) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    ws_ru = wb.create_sheet("DASHBOARD_TOP")
    _create_dashboard_top(ws_ru, totals, stats, lang="ru")

    ws_uz = wb.create_sheet("DASHBOARD_TOP_UZ")
    _create_dashboard_top(ws_uz, totals, stats, lang="uz")

    _create_all_pretty(wb, totals, stats)

    tmp_dir = tempfile.gettempdir()
    path = os.path.join(tmp_dir, "survey_stats.xlsx")
    wb.save(path)
    return path
